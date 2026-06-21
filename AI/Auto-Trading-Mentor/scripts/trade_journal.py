#!/usr/bin/env python3
"""Append AI trading run records to daily CSV/XLSX files and rebuild reports."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - environment dependent
    Workbook = None
    load_workbook = None


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MARKET = "mnq-demo"

HEADERS = [
    "run_id",
    "timestamp",
    "trading_date",
    "account_mode",
    "account_id",
    "symbol",
    "position",
    "position_size",
    "avg_price",
    "qty",
    "current_price",
    "open_pnl",
    "realized_pnl",
    "daily_pnl",
    "equity",
    "working_orders",
    "action_taken",
    "execution_decision",
    "fill_status",
    "direction",
    "entry",
    "stop",
    "take_profit",
    "invalidation",
    "rr",
    "setup_score",
    "confidence",
    "risk_status",
    "hard_blocks",
    "soft_warnings",
    "manual_intervention_required",
    "agent_risk_manager",
    "agent_1m_trader",
    "agent_5m_trader",
    "agent_1h_trader",
    "agent_support_resistance",
    "agent_liquidity",
    "agent_volume",
    "final_recommendation",
    "next_step",
    "source_message",
]


SAMPLE_RECORD = {
    "timestamp": "2026-06-10T17:30:00+08:00",
    "account_mode": "Demo",
    "symbol": "MNQM6",
    "position": 0,
    "qty": 3,
    "current_price": 21950.25,
    "working_orders": 0,
    "execution_decision": "等待",
    "action_taken": "none",
    "fill_status": "not_submitted",
    "candidate_setup": {
        "direction": "long",
        "entry": 21952.0,
        "stop": 21927.0,
        "take_profit": 21990.0,
        "invalidation": "跌破 21927 並失守 VWAP",
        "rr": 1.52,
        "setup_score": 69,
        "confidence": 64,
    },
    "risk": {
        "status": "soft_block",
        "hard_blocks": [],
        "soft_warnings": ["5m structure unavailable", "MACD unavailable"],
    },
    "agents": {
        "risk_manager": "未達 confidence 門檻，等待更清楚回測。",
        "one_min_trader": "1m 仍偏多，但進場點接近壓力。",
        "five_min_trader": "5m 結構不可見，降低分數。",
        "one_hour_trader": "1h 趨勢不可見。",
        "support_resistance": "支撐 21927，壓力 21990。",
        "liquidity": "上方流動性可能在 21990 附近。",
        "volume": "量能不足以確認突破。",
    },
    "final_recommendation": "等待",
    "next_step": "等待突破後回測，重新評分。",
    "source_message": "heartbeat XML message can be copied here",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--base-dir",
        default=str(REPO_ROOT),
        help="Project base directory. Defaults to the repository root.",
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    log_parser = subparsers.add_parser("log", help="Append one run record and rebuild the daily report.")
    log_parser.add_argument(
        "--base-dir",
        default=argparse.SUPPRESS,
        help="Project base directory. Can be placed before or after the log subcommand.",
    )
    log_parser.add_argument("--input", "-i", help="JSON record file. Use '-' for stdin.")
    log_parser.add_argument("--message", help="Raw heartbeat/message text to store when no JSON file is available.")
    log_parser.add_argument("--date", help="Trading date override in YYYY-MM-DD.")
    log_parser.add_argument("--market", default=DEFAULT_MARKET, help="Filename market prefix.")
    log_parser.add_argument("--no-xlsx", action="store_true", help="Skip XLSX output.")

    report_parser = subparsers.add_parser("report", help="Rebuild a daily report from an existing CSV.")
    report_parser.add_argument(
        "--base-dir",
        default=argparse.SUPPRESS,
        help="Project base directory. Can be placed before or after the report subcommand.",
    )
    report_parser.add_argument("--date", required=True, help="Trading date in YYYY-MM-DD.")
    report_parser.add_argument("--market", default=DEFAULT_MARKET, help="Filename market prefix.")

    sample_parser = subparsers.add_parser("sample", help="Print a sample JSON record.")
    sample_parser.add_argument("--compact", action="store_true", help="Print compact JSON.")

    return parser.parse_args()


def read_json_record(input_path: str | None, message: str | None) -> dict[str, Any]:
    if input_path:
        raw = sys.stdin.read() if input_path == "-" else Path(input_path).read_text(encoding="utf-8")
        try:
            return json.loads(raw)
        except json.JSONDecodeError as exc:
            raise SystemExit(f"Invalid JSON input: {exc}") from exc

    if message:
        return {"source_message": message}

    raise SystemExit("Provide --input JSON or --message text.")


def nested_get(data: dict[str, Any], path: str) -> Any:
    current: Any = data
    for part in path.split("."):
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def first_value(data: dict[str, Any], *paths: str) -> Any:
    for path in paths:
        value = nested_get(data, path) if "." in path else data.get(path)
        if value is not None and value != "":
            return value
    return ""


def scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, tuple, set)):
        return "; ".join(scalar(item) for item in value if item is not None)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def parse_timestamp(value: Any) -> datetime:
    text = scalar(value)
    if not text:
        return datetime.now().astimezone()
    normalized = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        return datetime.now().astimezone()


def normalize_record(data: dict[str, Any], trading_date: str | None) -> dict[str, str]:
    timestamp = parse_timestamp(first_value(data, "timestamp", "current_time_iso", "time"))
    date_text = trading_date or first_value(data, "trading_date", "date") or timestamp.date().isoformat()
    run_id = first_value(data, "run_id", "id")
    if not run_id:
        run_id = f"{date_text}-{timestamp.strftime('%H%M%S')}"

    agents = first_value(data, "agents", "agent_views")
    if not isinstance(agents, dict):
        agents = {}

    normalized = {
        "run_id": scalar(run_id),
        "timestamp": timestamp.isoformat(),
        "trading_date": scalar(date_text),
        "account_mode": scalar(first_value(data, "account_mode", "environment_status.account_mode", "authorization_mode.mode")),
        "account_id": scalar(first_value(data, "account_id", "environment_status.account_id")),
        "symbol": scalar(first_value(data, "symbol", "market", "market_scan_result.symbol", "environment_status.symbol")),
        "position": scalar(first_value(data, "position", "position_result.position")),
        "position_size": scalar(first_value(data, "position_size", "position_result.size")),
        "avg_price": scalar(first_value(data, "avg_price", "position_result.avg_price")),
        "qty": scalar(first_value(data, "qty", "quantity", "execution_decision.qty")),
        "current_price": scalar(first_value(data, "current_price", "market_scan_result.current_price")),
        "open_pnl": scalar(first_value(data, "open_pnl", "position_result.open_pnl")),
        "realized_pnl": scalar(first_value(data, "realized_pnl", "position_result.realized_pnl")),
        "daily_pnl": scalar(first_value(data, "daily_pnl", "risk_result.daily_pnl")),
        "equity": scalar(first_value(data, "equity", "risk_result.equity")),
        "working_orders": scalar(first_value(data, "working_orders", "orders", "position_result.working_orders")),
        "action_taken": scalar(first_value(data, "action_taken", "execution.action_taken")),
        "execution_decision": scalar(first_value(data, "execution_decision", "final", "decision", "execution_decision.final")),
        "fill_status": scalar(first_value(data, "fill_status", "execution.fill_status")),
        "direction": scalar(first_value(data, "direction", "candidate_setup.direction", "setup.direction")),
        "entry": scalar(first_value(data, "entry", "candidate_setup.entry", "setup.entry")),
        "stop": scalar(first_value(data, "stop", "candidate_setup.stop", "setup.stop")),
        "take_profit": scalar(first_value(data, "take_profit", "candidate_setup.take_profit", "setup.take_profit")),
        "invalidation": scalar(first_value(data, "invalidation", "candidate_setup.invalidation", "setup.invalidation")),
        "rr": scalar(first_value(data, "rr", "candidate_setup.rr", "setup.rr")),
        "setup_score": scalar(first_value(data, "setup_score", "candidate_setup.setup_score", "setup.setup_score")),
        "confidence": scalar(first_value(data, "confidence", "candidate_setup.confidence", "setup.confidence")),
        "risk_status": scalar(first_value(data, "risk_status", "risk.status", "risk_result.status")),
        "hard_blocks": scalar(first_value(data, "hard_blocks", "risk.hard_blocks", "safety_result.hard_blocks")),
        "soft_warnings": scalar(first_value(data, "soft_warnings", "risk.soft_warnings", "risk_result.soft_warnings")),
        "manual_intervention_required": scalar(
            first_value(data, "manual_intervention_required", "risk.manual_intervention_required")
        ),
        "agent_risk_manager": scalar(first_agent(agents, "risk_manager", "Risk Manager")),
        "agent_1m_trader": scalar(first_agent(agents, "one_min_trader", "1m_trader", "1m Trader")),
        "agent_5m_trader": scalar(first_agent(agents, "five_min_trader", "5m_trader", "5m Trader")),
        "agent_1h_trader": scalar(first_agent(agents, "one_hour_trader", "1h_trader", "1h Trader")),
        "agent_support_resistance": scalar(first_agent(agents, "support_resistance", "Support Resistance")),
        "agent_liquidity": scalar(first_agent(agents, "liquidity", "Liquidity")),
        "agent_volume": scalar(first_agent(agents, "volume", "Volume")),
        "final_recommendation": scalar(first_value(data, "final_recommendation", "final_suggestion", "message.final_recommendation")),
        "next_step": scalar(first_value(data, "next_step", "next_run_instruction", "message.next_step")),
        "source_message": scalar(first_value(data, "source_message", "message", "heartbeat_xml")),
    }
    return normalized


def first_agent(agents: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in agents:
            return agents[key]
    return ""


def slug(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip().lower())
    return cleaned.strip("-") or DEFAULT_MARKET


def output_paths(base_dir: Path, market: str, trading_date: str) -> dict[str, Path]:
    name = f"{trading_date}-{slug(market)}-trading-journal"
    journal_dir = base_dir / "logs" / "journal"
    report_dir = base_dir / "logs" / "reports"
    return {
        "csv": journal_dir / f"{name}.csv",
        "xlsx": journal_dir / f"{name}.xlsx",
        "report": report_dir / f"{trading_date}-{slug(market)}-trading-report.md",
    }


def append_csv(path: Path, record: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    exists = path.exists() and path.stat().st_size > 0
    with path.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        if not exists:
            writer.writeheader()
        writer.writerow({header: record.get(header, "") for header in HEADERS})


def sync_xlsx(csv_path: Path, xlsx_path: Path) -> bool:
    if Workbook is None or load_workbook is None:
        return False

    with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Journal"
    sheet.append(HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, header in enumerate(HEADERS, start=1):
        values = [header] + [row.get(header, "") for row in rows]
        width = min(max(len(str(value)) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(xlsx_path)
    return True


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise SystemExit(f"CSV not found: {path}")
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def to_float(value: str) -> float | None:
    if value is None:
        return None
    text = str(value).replace(",", "").replace("$", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def count_non_empty(rows: list[dict[str, str]], key: str) -> int:
    return sum(1 for row in rows if row.get(key))


def build_report(rows: list[dict[str, str]], trading_date: str, market: str) -> str:
    action_counts = Counter(row.get("action_taken", "") or "none" for row in rows)
    decision_counts = Counter(row.get("execution_decision", "") or "unknown" for row in rows)
    symbol_counts = Counter(row.get("symbol", "") or "unknown" for row in rows)
    pnl_values = [value for value in (to_float(row.get("daily_pnl", "")) for row in rows) if value is not None]
    score_values = [value for value in (to_float(row.get("setup_score", "")) for row in rows) if value is not None]
    confidence_values = [value for value in (to_float(row.get("confidence", "")) for row in rows) if value is not None]
    hard_block_count = count_non_empty(rows, "hard_blocks")
    manual_count = sum(1 for row in rows if row.get("manual_intervention_required", "").lower() == "true")

    latest_pnl = pnl_values[-1] if pnl_values else None
    max_drawdown = min(pnl_values) if pnl_values else None
    best_score = max(score_values) if score_values else None
    best_confidence = max(confidence_values) if confidence_values else None

    lines = [
        f"# {trading_date} {market.upper()} Trading Report",
        "",
        "## Summary",
        "",
        f"- Total runs: {len(rows)}",
        f"- Symbols: {format_counter(symbol_counts)}",
        f"- Decisions: {format_counter(decision_counts)}",
        f"- Actions: {format_counter(action_counts)}",
        f"- Latest daily PnL: {format_number(latest_pnl)}",
        f"- Worst recorded daily PnL: {format_number(max_drawdown)}",
        f"- Best setup score: {format_number(best_score)}",
        f"- Best confidence: {format_number(best_confidence)}",
        f"- Runs with hard blocks: {hard_block_count}",
        f"- Manual intervention flags: {manual_count}",
        "",
        "## Recent Runs",
        "",
        "| Time | Symbol | Decision | Action | Direction | Entry | Stop | TP | RR | Score | Confidence | Risk |",
        "| --- | --- | --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | --- |",
    ]

    for row in rows[-20:]:
        time_text = row.get("timestamp", "")
        if "T" in time_text:
            time_text = time_text.split("T", 1)[1][:8]
        cells = [
            time_text,
            row.get("symbol", ""),
            row.get("execution_decision", ""),
            row.get("action_taken", ""),
            row.get("direction", ""),
            row.get("entry", ""),
            row.get("stop", ""),
            row.get("take_profit", ""),
            row.get("rr", ""),
            row.get("setup_score", ""),
            row.get("confidence", ""),
            row.get("risk_status", ""),
        ]
        lines.append("| " + " | ".join(sanitize_md_cell(cell) for cell in cells) + " |")

    lines.extend(
        [
            "",
            "## Review Checklist",
            "",
            "- Compare high-score waiting setups against later price action.",
            "- Mark whether each hard block prevented a poor trade or missed an opportunity.",
            "- Review trades where confidence was high but RR or stop placement was weak.",
            "- Tighten prompt rules only after at least 20 comparable records.",
            "",
        ]
    )
    return "\n".join(lines)


def sanitize_md_cell(value: str) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def format_counter(counter: Counter[str]) -> str:
    if not counter:
        return "n/a"
    return ", ".join(f"{key}={value}" for key, value in counter.most_common())


def format_number(value: float | None) -> str:
    if value is None:
        return "n/a"
    if value.is_integer():
        return str(int(value))
    return f"{value:.2f}"


def write_report(path: Path, rows: list[dict[str, str]], trading_date: str, market: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(build_report(rows, trading_date, market), encoding="utf-8")


def command_log(args: argparse.Namespace) -> None:
    base_dir = Path(args.base_dir).resolve()
    raw_record = read_json_record(args.input, args.message)
    record = normalize_record(raw_record, args.date)
    paths = output_paths(base_dir, args.market, record["trading_date"])

    append_csv(paths["csv"], record)
    xlsx_written = False if args.no_xlsx else sync_xlsx(paths["csv"], paths["xlsx"])
    rows = read_csv(paths["csv"])
    write_report(paths["report"], rows, record["trading_date"], args.market)

    print(f"CSV: {paths['csv']}")
    if xlsx_written:
        print(f"XLSX: {paths['xlsx']}")
    elif not args.no_xlsx:
        print("XLSX: skipped (openpyxl is not installed)")
    print(f"REPORT: {paths['report']}")


def command_report(args: argparse.Namespace) -> None:
    base_dir = Path(args.base_dir).resolve()
    paths = output_paths(base_dir, args.market, args.date)
    rows = read_csv(paths["csv"])
    write_report(paths["report"], rows, args.date, args.market)
    print(f"REPORT: {paths['report']}")


def command_sample(args: argparse.Namespace) -> None:
    if args.compact:
        print(json.dumps(SAMPLE_RECORD, ensure_ascii=False, separators=(",", ":")))
    else:
        print(json.dumps(SAMPLE_RECORD, ensure_ascii=False, indent=2))


def main() -> None:
    args = parse_args()
    if args.command == "log":
        command_log(args)
    elif args.command == "report":
        command_report(args)
    elif args.command == "sample":
        command_sample(args)
    else:  # pragma: no cover
        raise SystemExit(f"Unknown command: {args.command}")


if __name__ == "__main__":
    main()
